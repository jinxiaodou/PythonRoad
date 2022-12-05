#!usr/bin/python
#coding=utf-8

import requests
from pyquery import PyQuery as pq
import openpyxl
import sys
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import io
import PIL # 需安装pip3 install Pillow

def send_request(url):
	headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36'}
	resp=requests.get(url, headers=headers)
	return resp

def get_cate(doc):
	# 获取分类数据
	catelist=doc('.homepage-cat-name').items()
	create_excel()
	for cate in catelist:
		href=url+cate.attr('href')
		catename=cate.text()
		cate_doc=pq(send_request(href).text)
		foodsList=get_food(cate_doc)
		if len(foodsList)>0:
			write_excel(foodsList,catename)
	delete_Sheet()

def get_food(cate_doc):
	# 获取美食数据
	foods=cate_doc('.category-recipe-list .list li').items()
	foodsList=[]
	for food in foods:
		name=food('.name a').text()
		s=food('.ing.ellipsis').text()
		scores=food('.stats span')
		score1=scores[0].text
		score2=scores[1].text if len(scores)>1 else '0'
		img=food('.cover.pure-u img').attr('data-src')
		href=url+food('div>a').attr('href')
		foodsList.append([name,s,score1,score2,img,href])
	return foodsList

def create_excel():
	# 创建excel
	wb=openpyxl.Workbook()
	wb.save('下厨房.xlsx')

def delete_Sheet():
	# 删除默认sheet
	wb=openpyxl.load_workbook('下厨房.xlsx')
	del wb['Sheet']
	wb.save('下厨房.xlsx')

def write_excel(list, sheetname):
	# 数据写入表格
	wb=openpyxl.load_workbook('下厨房.xlsx')
	sheet=wb.create_sheet(sheetname)
	sheet.append(['菜名','材料','评分','人气','图片','链接'])
	# 设置宽高 设置属性必须在设置数据之后，否则会独立成行
	set_styles(sheet)
	count=1
	for food in list:
		sheet.append(food)
		count+=1
		sheet.row_dimensions[count].height=120
		cell_num='E'+str(count)
		sheet[cell_num]=''			# 清空图片url
		# 插入图片
		resp=send_request(food[4])
		image_file = io.BytesIO(resp.content)
		img = Image(image_file)
		sheet.add_image(img,cell_num)
	wb.save('下厨房.xlsx')

# 整行整列设置样式失败
def set_styles(sheet):
	sheet.column_dimensions['A'].width=40
	sheet.column_dimensions['B'].width=60
	sheet.column_dimensions['E'].width=50
	sheet.column_dimensions['F'].width=50
	# sheet.column_dimensions['A'].font=Font(size=14)
	# sheet.column_dimensions['B'].font=Font(size=14)
	# sheet.column_dimensions['C'].font=Font(size=14)
	# sheet.column_dimensions['D'].font=Font(size=14)
	# sheet.column_dimensions['F'].font=Font(size=14)
	# sheet.row_dimensions[1].font=Font(size=16,bold=True)
	sheet.row_dimensions[1].height=20
	# 设置对齐方式
	# alignment=Alignment(
	# 	horizontal='left',
	# 	vertical='center'
	# 	)
	# alignment_center=Alignment(
	# 	horizontal='center',
	# 	vertical='center'
	# 	)
	# sheet.column_dimensions['A'].alignment=alignment
	# sheet.column_dimensions['B'].alignment=alignment
	# sheet.column_dimensions['C'].alignment=alignment
	# sheet.column_dimensions['D'].alignment=alignment
	# sheet.column_dimensions['E'].alignment=alignment
	# sheet.column_dimensions['F'].alignment=alignment
	# sheet.row_dimensions[1].alignment=alignment_center

def start():
	doc=pq(send_request(url).text)
	get_cate(doc)

url='https://www.xiachufang.com'
start()
