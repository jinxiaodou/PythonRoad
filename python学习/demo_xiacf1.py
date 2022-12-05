from pyquery import PyQuery as pq
import openpyxl
import sys
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import io
import PIL # 需安装pip3 install Pillow

def set 

def set_styles(cell):
	# sheet.column_dimensions['A'].width=40
	# sheet.column_dimensions['B'].width=60
	# sheet.column_dimensions['E'].width=50
	# sheet.column_dimensions['F'].width=50
	font14=Font(size=14)
	font16=Font(size=16,bold=True)
	cell.font=font14
	# sheet.column_dimensions['B'].font=font14
	# sheet.column_dimensions['C'].font=font14
	# sheet.column_dimensions['D'].font=font14
	# sheet.column_dimensions['F'].font=font14
	# sheet.row_dimensions[1].font=font16
	# sheet.row_dimensions[1].height=20
	# 设置对齐方式
	alignment=Alignment(
		horizontal='left',
		vertical='center'
		)
	# alignment_center=Alignment(
	# 	horizontal='center',
	# 	vertical='center'
	# 	)
	cell.alignment=alignment
	# sheet.column_dimensions['B'].alignment=alignment
	# sheet.column_dimensions['C'].alignment=alignment
	# sheet.column_dimensions['D'].alignment=alignment
	# sheet.column_dimensions['E'].alignment=alignment
	# sheet.column_dimensions['F'].alignment=alignment
	# sheet.row_dimensions[1].alignment=alignment_center


def set_Sheet():
	# 删除默认sheet
	wb=openpyxl.load_workbook('xxx.xlsx')
	# wb.save('xxx.xlsx')
	for sheet_name in wb.sheetnames:
		sheet=wb[sheet_name]
		for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row):
			for cell in row:
				set_styles(cell)
	wb.save('xxx.xlsx')

set_Sheet()