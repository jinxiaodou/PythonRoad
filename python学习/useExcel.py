#!usr/bin/python
#coding=utf-8

# 非内嵌模块，需pip导入
import openpyxl
import requests


def send_request(url):
	headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36'}
	resp=requests.get(url, headers=headers)
	return resp

# Excel写入数据
# 创建工作簿对象
wb=openpyxl.Workbook()
# 获取工作表对象
sheet=wb.active
# 获取指定单元格
cell=sheet['A1']
# 向单元格中写入数据
cell.value='一步之遥'
# 写入一行数据
list=['姓名','年龄','职位']
# 设置宽高 仅设置一个
sheet.column_dimensions['A'].width=40
sheet.row_dimensions[1].height=20
sheet.row_dimensions[2].height=30
sheet.append(list)
# 设置对齐方式 只能设置单元格，且需存在再设置，否则append会向下添加一行
sheet['B3'].alignment=openpyxl.styles.Alignment(
		horizontal='left',
		vertical='center'
		)
# 写入多行数据
list1=[
	['姜文',35,'导演'],
	['葛优',40,'副导演']
]
count=2
for item in list1:
	count+=1
	sheet.append(item)
	sheet.row_dimensions[count].height=40
# 写入新sheet
sheet1=wb.create_sheet()
sheet1.title='演员表'

# 保存
wb.save('一步之遥.xlsx')

# Excel读取数据
# 加载Excel文件
wb1=openpyxl.load_workbook('一步之遥.xlsx')
# 获取指定sheet
sheet1=wb['Sheet']
print('获取A列')
colums=sheet1['A']
for cell in colums:
	print(cell.value)
print('获取第3行')
rows=sheet1[3]
for cell in rows:
	print(cell.value)
print('获取BC列')
cols=sheet1['B:C']
for col in cols:
	for cell in col:
		print(cell.value)


